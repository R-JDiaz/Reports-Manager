from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from format import autofit_selected_columns

def generateDefaultFormat(filename="sample2"):
    wb = Workbook()

    sheet1 = wb.active
    sheet1.title = "General"
    sheet2 = wb.create_sheet(title="Doctors")
    sheet3 = wb.create_sheet(title="Exams")

    sheet1['A1'] = "Patient's Name"
    sheet1['B1'] = "Exam Type"
    sheet1['C1'] = "Doctor's Name"
    autofit_selected_columns(sheet1, [1,2,3])

    sheet2['A1'] = "Doctor's Name"
    sheet2['B1'] = "Email"
    autofit_selected_columns(sheet2, [1,2])

    sheet3['A1'] = "Exam Type"
    sheet3['B1'] = "Commission Amount"
    autofit_selected_columns(sheet3, [1,2])

    opt1 = DefinedName(
        name="doctorsName", 
        attr_text="OFFSET(Doctors!$A$2, 0, 0, COUNTA(Doctors!$A:$A)-1, 1)")

    opt2 = DefinedName(
        name="examNames",
        attr_text="OFFSET(Exams!$A$2, 0, 0, COUNTA(Exams!$A:$A)-1, 1)")

    dvDoctors = DataValidation(type="list", formula1="=doctorsName", allow_blank=True)
    dvExams = DataValidation(type="list", formula1="=examNames", allow_blank=True)

    sheet1.add_data_validation(dvDoctors)
    sheet1.add_data_validation(dvExams)
    dvExams.add("B2:B300")
    dvDoctors.add("C2:C300")

    wb.defined_names.add(opt1)
    wb.defined_names.add(opt2)

    wb.save(f"./excel/{filename}.xlsx")