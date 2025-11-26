from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

font = Font(
    name='Arial',
    size=11,
    bold=True
)

def autofit_selected_columns(ws, column):
    for num in column:
        cellRef = get_column_letter(num)
        cell = ws[f"{cellRef}1"]
        length = len(str(cell.value)) + 4
        ws.column_dimensions[get_column_letter(num)].width = length
        setFont(cell)

def setFont(cell, myFont=font):
    cell.font = myFont
