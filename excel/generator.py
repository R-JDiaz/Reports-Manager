from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from copy import copy


def copy_excel_structure(src_path, dst_path, copy_values=False):
    print("Loading source workbook...")
    src_wb = load_workbook(src_path, data_only=False)
    dst_wb = Workbook()

    # Remove the default sheet
    if "Sheet" in dst_wb.sheetnames:
        del dst_wb["Sheet"]

    print("Copying sheets...")
    for sheet_name in src_wb.sheetnames:
        src_ws = src_wb[sheet_name]
        dst_ws = dst_wb.create_sheet(sheet_name)

        print(f" - Copying: {sheet_name}")

        # ---- Copy sheet-level properties ----
        dst_ws.sheet_properties.outlinePr = copy(src_ws.sheet_properties.outlinePr)
        dst_ws.sheet_properties.tabColor = src_ws.sheet_properties.tabColor

        # ---- Copy page setup ----
        dst_ws.page_setup.orientation = src_ws.page_setup.orientation
        dst_ws.page_setup.paperSize = src_ws.page_setup.paperSize
        dst_ws.page_setup.fitToWidth = src_ws.page_setup.fitToWidth
        dst_ws.page_setup.fitToHeight = src_ws.page_setup.fitToHeight

        dst_ws.page_margins.left = src_ws.page_margins.left
        dst_ws.page_margins.right = src_ws.page_margins.right
        dst_ws.page_margins.top = src_ws.page_margins.top
        dst_ws.page_margins.bottom = src_ws.page_margins.bottom

        # ---- Copy print options ----
        dst_ws.print_options.horizontalCentered = src_ws.print_options.horizontalCentered
        dst_ws.print_options.verticalCentered = src_ws.print_options.verticalCentered

        # ---- Copy row & column dimensions ----
        for col_letter, dim in src_ws.column_dimensions.items():
            dst_ws.column_dimensions[col_letter].width = dim.width

        for row_idx, dim in src_ws.row_dimensions.items():
            dst_ws.row_dimensions[row_idx].height = dim.height

        # ---- Copy merged cells ----
        for merged in src_ws.merged_cells.ranges:
            dst_ws.merge_cells(str(merged))

        # ---- Copy cells ----
        for row in src_ws.iter_rows():
            for cell in row:
                new_cell = dst_ws[cell.coordinate]

                # Value / Formula
                if copy_values:
                    new_cell.value = cell.value
                else:
                    if cell.data_type == "f":  # keep formulas
                        new_cell.value = cell.value

                # Styles
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

        # ---- Copy Data Validations ----
        print(f"   - Copying Data Validations")
        for dv in src_ws.data_validations.dataValidation:
            new_dv = DataValidation(
                type=dv.type,
                formula1=dv.formula1,
                formula2=dv.formula2,
                allow_blank=dv.allow_blank,
                showDropDown=dv.showDropDown,
                promptTitle=dv.promptTitle,
                prompt=dv.prompt,
                errorTitle=dv.errorTitle,
                error=dv.error
            )
            dst_ws.add_data_validation(new_dv)

            # Copy all ranges inside DV
            for rng in dv.cells.ranges:
                new_dv.add(rng)

    # ---- Copy Named Ranges (Workbook Level) ----
    print("Copying named ranges...")
    for name in src_wb.defined_names.definedName:
        dst_wb.defined_names.append(name)

    print("Saving workbook...")
    dst_wb.save(dst_path)
    print("Done! →", dst_path)


# ---- RUN ----
copy_excel_structure("format.xlsx", "copy.xlsx", copy_values=False)
